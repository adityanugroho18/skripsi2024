from django.shortcuts import render, redirect, get_object_or_404
from dashboard.utils import date_format
from log_activity.models import LogActivity, LogActivityDetail
import os
from django.conf import settings
from django.contrib import messages
from dashboard.models import Jurnal
from django.db.models import Q, Sum
from django.db import transaction
from datetime import datetime
from django.http import HttpResponse
from kode_akun.models import KodeAkun
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from .models import ViewLabaRugi

# Create your views here.
context = {
    'app_name': os.environ.get('APP_NAME', settings.APP_NAME)
}

# Buku Besar
def index(request):
    context['title'] = 'Buku Besar'
    context['kode_akun'] = KodeAkun.objects.all()
    form = request.GET
    context['fieldValues'] = form
    kode_akun = form.getlist('kode_akun', None)
    dari = form.get('dari', None)
    sampai = form.get('sampai', None)
    dari_formated = (date_format(dari, '%Y-%m-%d', '%d-%m-%Y') if dari != None else dari)
    sampai_formated = (date_format(sampai, '%Y-%m-%d', '%d-%m-%Y') if sampai != None else sampai)
    context['dari'] = dari_formated
    context['sampai'] = sampai_formated
    context['query_params'] = None

    if len(kode_akun) > 0 and dari and sampai:
        # Get the query parameters as a dictionary
        query_params = '?'
        if len(kode_akun) > 0:
            for kode in kode_akun:
                query_params += f'kode_akun={kode}&'
        query_params += f'dari={dari}&sampai={sampai}'

        context['query_params'] = query_params

        # Handle 'all' case if necessary
        if 'all' in kode_akun:
            kode_akun = kode_akun[0]

        if isinstance(kode_akun, list):
            selected_kode_akun = KodeAkun.objects.filter(kode_akun__in=kode_akun)
        else:
            selected_kode_akun = KodeAkun.objects.all()
        context['option_selected_kode_akun'] = kode_akun
        
        context['selected_kode_akun'] = selected_kode_akun

        for item in selected_kode_akun:
            total_debit = 0
            total_kredit = 0
            saldo_akhir = 0
            saldo_awal_debit2 = 0
            saldo_awal_kredit2 = 0
            saldo_awal_debit = 0
            saldo_awal_kredit = 0

            # count jumlah transaksi masing2 kode rekening sebelum tanggal dari
            transaksi_exists = Jurnal.objects.filter(
                Q(kode_akun=item.kode_akun) | Q(kode_lawan=item.kode_akun)
            ).exists()
            
            if transaksi_exists:
                jurnal_exists = Jurnal.objects.filter(
                    kode_akun=item.kode_akun,
                    tanggal__lt=dari
                ).exists()
                
                if jurnal_exists:
                    saldo_awal_debit = Jurnal.objects.filter(
                                            kode_akun=item.kode_akun,
                                            tipe='debit',
                                            tanggal__lt=dari
                                        ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                    saldo_awal_kredit = Jurnal.objects.filter(
                                            kode_akun=item.kode_akun,
                                            tipe='kredit',
                                            tanggal__lt=dari
                                        ).aggregate(total_nominal=Sum('nominal'))['total_nominal']

                    # cek apakah rekening juga terdapat di field lawan di table jurnal
                    lawan_exists = Jurnal.objects.filter(
                        kode_lawan=item.kode_akun,
                        tanggal__lt=dari
                    ).exists()
                    if lawan_exists:
                        saldo_awal_debit2 = Jurnal.objects.filter(
                                                kode_akun=item.kode_akun,
                                                tipe='kredit',
                                                tanggal__lt=dari
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                        saldo_awal_kredit2 = Jurnal.objects.filter(
                                                kode_akun=item.kode_akun,
                                                tipe='debit',
                                                tanggal__lt=dari
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                else:
                    # rekening tsb tidak terdapat di field kode dan hanya terdapat di field lawan
                    saldo_awal_debit = Jurnal.objects.filter(
                                            kode_akun=item.kode_akun,
                                            tipe='kredit',
                                            tanggal__lt=dari
                                        ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                    saldo_awal_debit = Jurnal.objects.filter(
                                            kode_akun=item.kode_akun,
                                            tipe='debit',
                                            tanggal__lt=dari
                                        ).aggregate(total_nominal=Sum('nominal'))['total_nominal']

                # set to 0 when returned None
                saldo_awal_debit2 = (0 if not saldo_awal_debit2 else saldo_awal_debit2)
                saldo_awal_kredit2 = (0 if not saldo_awal_kredit2 else saldo_awal_kredit2)
                saldo_awal_debit = (0 if not saldo_awal_debit else saldo_awal_debit)
                saldo_awal_kredit = (0 if not saldo_awal_kredit else saldo_awal_kredit)
                # hitung saldoAwal dari rekening
                if item.tipe == 'debit':
                    saldo_akhir = (saldo_awal_debit + saldo_awal_debit2) - (saldo_awal_kredit + saldo_awal_kredit2);
                else:
                    saldo_akhir = (saldo_awal_debit + saldo_awal_debit2) + (saldo_awal_kredit + saldo_awal_kredit2);
            else:
                # set saldo akhir = saldo awal rekening
                saldo_akhir = 0

            buku_besar = Jurnal.objects.filter(
                                Q(tanggal__range=(dari, sampai)),
                                Q(kode_akun=item.kode_akun) | Q(kode_lawan=item.kode_akun)
                            ).order_by('tanggal')
            for gl in buku_besar:
                # cek posisi lawan (pasangan) ada di field kode atau di field lawan
                if gl.kode_akun == item.kode_akun:
                    field_lawan = 'lawan'
                else:
                    field_lawan = 'kode'
                # Add field_lawan to item
                if isinstance(gl, dict):
                    gl['field_lawan'] = field_lawan
                else:
                    setattr(gl, 'field_lawan', field_lawan)

                if field_lawan == 'lawan':
                    kode_lawan_parent = KodeAkun.objects.filter(kode_akun=gl.kode_lawan).first()
                else:
                    kode_lawan_parent = KodeAkun.objects.filter(kode_akun=gl.kode_akun).first()
                # Add kode_lawan_parent to item
                if isinstance(gl, dict):
                    gl['kode_lawan_parent'] = kode_lawan_parent
                else:
                    setattr(gl, 'kode_lawan_parent', kode_lawan_parent)

                saldo_akhir_gl = 0

                if field_lawan == 'lawan':
                    # jika tipe transaksi = debet
                    if gl.tipe == 'debit':
                        # totaldebet bertambah
                        total_debit += gl.nominal
                        
                        # jika tipe rekening = debet
                        if item.tipe == 'debit':
                            #saldo akhir rekening bertambah
                            saldo_akhir_gl += gl.nominal
                        else:
                            # saldo akhir rekening berkurang
                            saldo_akhir_gl -= gl.nominal
                    else:
                        # total kredit bertambah
                        total_kredit += gl.nominal
                        
                        # jika tipe rekening = debet
                        if item.tipe == 'debit':
                            #saldo akhir rekening berkurang
                            saldo_akhir_gl -= gl.nominal
                        else:
                            # saldo akhir rekening bertambah
                            saldo_akhir_gl += gl.nominal
                else:
                    # jika lawan terdapat di field kode
                    # jika tipe transaksi  = debet
                    if gl.tipe == 'debit':
                        # total kredit bertambah
                        total_kredit += gl.nominal

                        # jika tipe rekening = Debit
                        if item.tipe == 'debit':
                            # saldo akhir berkurang
                            saldo_akhir_gl -= gl.nominal
                        else:
                            # saldo akhir bertambah
                            saldo_akhir_gl += gl.nominal
                    else:
                        # total debet bertambah
                        total_debit += gl.nominal

                        # jika tipe rekening = Debit
                        if item.tipe == 'debit':
                            # saldo akhir bertambah
                            saldo_akhir_gl += gl.nominal
                        else:
                            # saldo akhir berkurang
                            saldo_akhir_gl -= gl.nominal
                
                if isinstance(gl, dict):
                    gl['saldo_akhir_gl'] = saldo_akhir_gl
                else:
                    setattr(gl, 'saldo_akhir_gl', saldo_akhir_gl)

            saldo_dict = {
                'total_debit': total_debit,
                'total_kredit': total_kredit,
                'saldo_akhir': saldo_akhir,
                'saldo_awal_debit2': saldo_awal_debit2,
                'saldo_awal_kredit2': saldo_awal_kredit2,
                'saldo_awal_debit': saldo_awal_debit,
                'saldo_awal_kredit': saldo_awal_kredit
            }
            # Add saldo_dict to item
            if isinstance(item, dict):
                item['saldo_dict'] = saldo_dict
            else:
                setattr(item, 'saldo_dict', saldo_dict)
            # Add buku_besar to item
            if isinstance(item, dict):
                item['buku_besar'] = buku_besar
            else:
                setattr(item, 'buku_besar', buku_besar)
    else:
        context['selected_kode_akun'] = []
        context['query_params'] = None
        context['option_selected_kode_akun'] = None

    return render(request, 'general-ledger/buku-besar.html', context)

def export_to_excel(request):
    form = request.GET
    kode_akun = form.getlist('kode_akun', None)
    dari = form.get('dari', None)
    sampai = form.get('sampai', None)
    dari_formated = (date_format(dari, '%Y-%m-%d', '%d-%m-%Y') if dari != None else dari)
    sampai_formated = (date_format(sampai, '%Y-%m-%d', '%d-%m-%Y') if sampai != None else sampai)

    if len(kode_akun) > 0 and dari and sampai:
        # Handle 'all' case if necessary
        if 'all' in kode_akun:
            kode_akun = kode_akun[0]

        if isinstance(kode_akun, list):
            selected_kode_akun = KodeAkun.objects.filter(kode_akun__in=kode_akun)
        else:
            selected_kode_akun = KodeAkun.objects.all()
        context['selected_kode_akun'] = selected_kode_akun

        for item in selected_kode_akun:
            total_debit = 0
            total_kredit = 0
            saldo_akhir = 0
            saldo_awal_debit2 = 0
            saldo_awal_kredit2 = 0
            saldo_awal_debit = 0
            saldo_awal_kredit = 0

            # count jumlah transaksi masing2 kode rekening sebelum tanggal dari
            transaksi_exists = Jurnal.objects.filter(
                Q(kode_akun=item.kode_akun) | Q(kode_lawan=item.kode_akun)
            ).exists()
            
            if transaksi_exists:
                jurnal_exists = Jurnal.objects.filter(
                    kode_akun=item.kode_akun,
                    tanggal__lt=dari
                ).exists()
                
                if jurnal_exists:
                    saldo_awal_debit = Jurnal.objects.filter(
                                            kode_akun=item.kode_akun,
                                            tipe='debit',
                                            tanggal__lt=dari
                                        ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                    saldo_awal_kredit = Jurnal.objects.filter(
                                            kode_akun=item.kode_akun,
                                            tipe='kredit',
                                            tanggal__lt=dari
                                        ).aggregate(total_nominal=Sum('nominal'))['total_nominal']

                    # cek apakah rekening juga terdapat di field lawan di table jurnal
                    lawan_exists = Jurnal.objects.filter(
                        kode_lawan=item.kode_akun,
                        tanggal__lt=dari
                    ).exists()
                    if lawan_exists:
                        saldo_awal_debit2 = Jurnal.objects.filter(
                                                kode_akun=item.kode_akun,
                                                tipe='kredit',
                                                tanggal__lt=dari
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                        saldo_awal_kredit2 = Jurnal.objects.filter(
                                                kode_akun=item.kode_akun,
                                                tipe='debit',
                                                tanggal__lt=dari
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                else:
                    # rekening tsb tidak terdapat di field kode dan hanya terdapat di field lawan
                    saldo_awal_debit = Jurnal.objects.filter(
                                            kode_akun=item.kode_akun,
                                            tipe='kredit',
                                            tanggal__lt=dari
                                        ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                    saldo_awal_debit = Jurnal.objects.filter(
                                            kode_akun=item.kode_akun,
                                            tipe='debit',
                                            tanggal__lt=dari
                                        ).aggregate(total_nominal=Sum('nominal'))['total_nominal']

                # set to 0 when returned None
                saldo_awal_debit2 = (0 if not saldo_awal_debit2 else saldo_awal_debit2)
                saldo_awal_kredit2 = (0 if not saldo_awal_kredit2 else saldo_awal_kredit2)
                saldo_awal_debit = (0 if not saldo_awal_debit else saldo_awal_debit)
                saldo_awal_kredit = (0 if not saldo_awal_kredit else saldo_awal_kredit)
                # hitung saldoAwal dari rekening
                if item.tipe == 'debit':
                    saldo_akhir = (saldo_awal_debit + saldo_awal_debit2) - (saldo_awal_kredit + saldo_awal_kredit2);
                else:
                    saldo_akhir = (saldo_awal_debit + saldo_awal_debit2) + (saldo_awal_kredit + saldo_awal_kredit2);
            else:
                # set saldo akhir = saldo awal rekening
                saldo_akhir = 0

            buku_besar = Jurnal.objects.filter(
                                Q(tanggal__range=(dari, sampai)),
                                Q(kode_akun=item.kode_akun) | Q(kode_lawan=item.kode_akun)
                            ).order_by('tanggal')
            for gl in buku_besar:
                # cek posisi lawan (pasangan) ada di field kode atau di field lawan
                if gl.kode_akun == item.kode_akun:
                    field_lawan = 'lawan'
                else:
                    field_lawan = 'kode'
                # Add field_lawan to item
                if isinstance(gl, dict):
                    gl['field_lawan'] = field_lawan
                else:
                    setattr(gl, 'field_lawan', field_lawan)

                if field_lawan == 'lawan':
                    kode_lawan_parent = KodeAkun.objects.filter(kode_akun=gl.kode_lawan).first()
                else:
                    kode_lawan_parent = KodeAkun.objects.filter(kode_akun=gl.kode_akun).first()
                # Add kode_lawan_parent to item
                if isinstance(gl, dict):
                    gl['kode_lawan_parent'] = kode_lawan_parent
                else:
                    setattr(gl, 'kode_lawan_parent', kode_lawan_parent)

                saldo_akhir_gl = 0

                if field_lawan == 'lawan':
                    # jika tipe transaksi = debet
                    if gl.tipe == 'debit':
                        # totaldebet bertambah
                        total_debit += gl.nominal
                        
                        # jika tipe rekening = debet
                        if item.tipe == 'debit':
                            #saldo akhir rekening bertambah
                            saldo_akhir_gl += gl.nominal
                        else:
                            # saldo akhir rekening berkurang
                            saldo_akhir_gl -= gl.nominal
                    else:
                        # total kredit bertambah
                        total_kredit += gl.nominal
                        
                        # jika tipe rekening = debet
                        if item.tipe == 'debit':
                            #saldo akhir rekening berkurang
                            saldo_akhir_gl -= gl.nominal
                        else:
                            # saldo akhir rekening bertambah
                            saldo_akhir_gl += gl.nominal
                else:
                    # jika lawan terdapat di field kode
                    # jika tipe transaksi  = debet
                    if gl.tipe == 'debit':
                        # total kredit bertambah
                        total_kredit += gl.nominal

                        # jika tipe rekening = Debit
                        if item.tipe == 'debit':
                            # saldo akhir berkurang
                            saldo_akhir_gl -= gl.nominal
                        else:
                            # saldo akhir bertambah
                            saldo_akhir_gl += gl.nominal
                    else:
                        # total debet bertambah
                        total_debit += gl.nominal

                        # jika tipe rekening = Debit
                        if item.tipe == 'debit':
                            # saldo akhir bertambah
                            saldo_akhir_gl += gl.nominal
                        else:
                            # saldo akhir berkurang
                            saldo_akhir_gl -= gl.nominal
                
                if isinstance(gl, dict):
                    gl['saldo_akhir_gl'] = saldo_akhir_gl
                else:
                    setattr(gl, 'saldo_akhir_gl', saldo_akhir_gl)

            saldo_dict = {
                'total_debit': total_debit,
                'total_kredit': total_kredit,
                'saldo_akhir': saldo_akhir,
                'saldo_awal_debit2': saldo_awal_debit2,
                'saldo_awal_kredit2': saldo_awal_kredit2,
                'saldo_awal_debit': saldo_awal_debit,
                'saldo_awal_kredit': saldo_awal_kredit
            }
            # Add saldo_dict to item
            if isinstance(item, dict):
                item['saldo_dict'] = saldo_dict
            else:
                setattr(item, 'saldo_dict', saldo_dict)
            # Add buku_besar to item
            if isinstance(item, dict):
                item['buku_besar'] = buku_besar
            else:
                setattr(item, 'buku_besar', buku_besar)

    wb = openpyxl.Workbook()
    ws = wb.active
    filename = f"Buku Besar Periode {dari} sd {sampai}"
    ws.title = filename

    # Define the header
    header = [
        "Tanggal", "Kode Transaksi", "Keterangan", "Lawan", 
        "Debit", "Kredit", "Saldo"
    ]

    # Append data rows
    for selected in selected_kode_akun:
        # Create section header row
        section_header = [f"{selected.nama} - {selected.kode_akun}", "", "", "", "", "", ""]
        ws.append(section_header)
        for cell in ws[ws.max_row]:  # Apply styles to the section header row
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Create header row
        ws.append(header)
        for cell in ws[ws.max_row]:  # Apply styles to the header row
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Add starting balance row
        ws.append([
            dari_formated, "-", "Saldo Awal", "", "", "", selected.saldo_dict['saldo_akhir']
        ])

        for gl in selected.buku_besar:
            ws.append([
                gl.tanggal.strftime('%d-%m-%Y'), gl.kode_transaksi, gl.keterangan or "-",
                f"{gl.kode_lawan} - {gl.kode_lawan_parent.nama}",
                gl.nominal if gl.tipe == 'debit' else "-",
                "-" if gl.tipe == 'debit' else gl.nominal,
                gl.saldo_akhir_gl
            ])

        # Add totals row
        ws.append([
            "", "", "", "Total", 
            selected.saldo_dict['total_debit'], 
            selected.saldo_dict['total_kredit'], ""
        ])

    # Create an HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
    wb.save(response)
    return response
# END Buku Besar

# Neraca Saldo
def index_neraca(request):
    context['title'] = 'Neraca Saldo'
    context['kode_akun'] = KodeAkun.objects.all()
    form = request.GET
    context['fieldValues'] = form
    kode_akun = form.getlist('kode_akun', None)
    dari = form.get('dari', None)
    sampai = form.get('sampai', None)
    dari_formated = (date_format(dari, '%Y-%m-%d', '%d-%m-%Y') if dari != None else dari)
    sampai_formated = (date_format(sampai, '%Y-%m-%d', '%d-%m-%Y') if sampai != None else sampai)
    context['dari'] = dari_formated
    context['sampai'] = sampai_formated
    context['query_params'] = None

    if len(kode_akun) > 0 and dari and sampai:
        # Get the query parameters as a dictionary
        query_params = '?'
        if len(kode_akun) > 0:
            for kode in kode_akun:
                query_params += f'kode_akun={kode}&'
        query_params += f'dari={dari}&sampai={sampai}'

        context['query_params'] = query_params

        # Handle 'all' case if necessary
        if 'all' in kode_akun:
            kode_akun = kode_akun[0]

        if isinstance(kode_akun, list):
            selected_kode_akun = KodeAkun.objects.filter(kode_akun__in=kode_akun)
        else:
            selected_kode_akun = KodeAkun.objects.all()
        context['option_selected_kode_akun'] = kode_akun
        context['selected_kode_akun'] = selected_kode_akun
        total_saldo_awal_debet = 0
        total_saldo_awal_kredit = 0
        total_mutasi_debet = 0
        total_mutasi_kredit = 0
        total_saldo_akhir_debet = 0
        total_saldo_akhir_kredit = 0

        for item in selected_kode_akun:
            mutasi_awal_debet = 0
            mutasi_awal_kredit = 0

            mutasi_debet = 0
            mutasi_kredit = 0

            # cek apakah ada jurnal awal di field kode
            cek_transaksi_awal_di_kode = Jurnal.objects.filter(
                kode_akun=item.kode_akun,
                tanggal__lt=dari
            ).exists()
            
            if cek_transaksi_awal_di_kode:
                sum_mutasi_awal_debet_di_kode = Jurnal.objects.filter(
                    kode_akun=item.kode_akun,
                    tanggal__lt=dari,
                    tipe='debit'
                ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                sum_mutasi_awal_kredit_di_kode = Jurnal.objects.filter(
                    kode_akun=item.kode_akun,
                    tanggal__lt=dari,
                    tipe='kredit'
                ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                
                mutasi_awal_debet += sum_mutasi_awal_debet_di_kode
                mutasi_awal_kredit += sum_mutasi_awal_kredit_di_kode

                # cek apakah transaksi sebelumnya juga terdapat di field lawan
                cek_transaksi_lawan_di_lawan = Jurnal.objects.filter(
                                                kode_lawan=item.kode_akun,
                                                tanggal__lt=dari
                                            ).exists()

                if cek_transaksi_lawan_di_lawan:
                    sum_mutasi_awal_debet_di_lawan = Jurnal.objects.filter(
                                                        kode_lawan=item.kode_akun,
                                                        tanggal__lt=dari,
                                                        tipe='debit'
                                                    ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                    sum_mutasi_awal_kredit_di_lawan = Jurnal.objects.filter(
                                                        kode_lawan=item.kode_akun,
                                                        tanggal__lt=dari,
                                                        tipe='kredit'
                                                    ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                    mutasi_awal_debet += sum_mutasi_awal_debet_di_lawan
                    mutasi_awal_kredit += sum_mutasi_awal_kredit_di_lawan
            else:
                # cek apakah ada jurnal awal di field lawan
                cek_transaksi_awal_di_lawan = Jurnal.objects.filter(
                                                    kode_lawan=item.kode_akun,
                                                    tanggal__lt=dari
                                                ).exists()
                if cek_transaksi_awal_di_lawan:
                    sum_mutasi_awal_debet_di_lawan = Jurnal.objects.filter(
                                                        kode_lawan=item.kode_akun,
                                                        tanggal__lt=dari,
                                                        tipe='kredit'
                                                    ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                    sum_mutasi_awal_kredit_di_lawan = Jurnal.objects.filter(
                                                        kode_lawan=item.kode_akun,
                                                        tanggal__lt=dari,
                                                        tipe='debit'
                                                    ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                    mutasi_awal_debet += sum_mutasi_awal_debet_di_lawan
                    mutasi_awal_kredit += sum_mutasi_awal_kredit_di_lawan

            # cek transaksi di field kode
            cek_transaksi_di_kode = Jurnal.objects.filter(
                Q(tanggal__range=(dari,sampai)),
                kode_akun=item.kode_akun
            ).exists()

            if cek_transaksi_di_kode:
                sum_mutasi_debet_di_kode = Jurnal.objects.filter(
                                            Q(tanggal__range=(dari,sampai)),
                                            kode_akun=item.kode_akun,
                                            tipe='debit'
                                        ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                sum_mutasi_kredit_di_kode = Jurnal.objects.filter(
                                            Q(tanggal__range=(dari,sampai)),
                                            kode_akun=item.kode_akun,
                                            tipe='kredit'
                                        ).aggregate(total_nominal=Sum('nominal'))['total_nominal']

                mutasi_debet += (0 if sum_mutasi_debet_di_kode == None else sum_mutasi_debet_di_kode)
                mutasi_kredit += (0 if sum_mutasi_kredit_di_kode == None else sum_mutasi_kredit_di_kode)

                # cek transaksi di field lawan
                cek_transaksi_di_lawan = Jurnal.objects.filter(
                                            Q(tanggal__range=(dari,sampai)),
                                            kode_lawan=item.kode_akun
                                        ).exists()
                if cek_transaksi_di_lawan:
                    sum_mutasi_debet_di_lawan = Jurnal.objects.filter(
                                            Q(tanggal__range=(dari,sampai)),
                                            kode_lawan=item.kode_akun,
                                            tipe='kredit'
                                        ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                    sum_mutasi_kredit_di_lawan = Jurnal.objects.filter(
                                                Q(tanggal__range=(dari,sampai)),
                                                kode_lawan=item.kode_akun,
                                                tipe='debit'
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']

                    mutasi_debet += (0 if sum_mutasi_debet_di_lawan == None else sum_mutasi_debet_di_lawan)
                    mutasi_kredit += (0 if sum_mutasi_kredit_di_lawan == None else sum_mutasi_kredit_di_lawan)
            else:
                # cek transaksi di field lawan
                cek_transaksi_di_lawan = Jurnal.objects.filter(
                                            Q(tanggal__range=(dari,sampai)),
                                            kode_lawan=item.kode_akun
                                        ).exists()
                if cek_transaksi_di_lawan:
                    sum_mutasi_debet_di_lawan = Jurnal.objects.filter(
                                            Q(tanggal__range=(dari,sampai)),
                                            kode_lawan=item.kode_akun,
                                            tipe='kredit'
                                        ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                    sum_mutasi_kredit_di_lawan = Jurnal.objects.filter(
                                                Q(tanggal__range=(dari,sampai)),
                                                kode_lawan=item.kode_akun,
                                                tipe='debit'
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']

                    mutasi_debet += (0 if sum_mutasi_debet_di_lawan == None else sum_mutasi_debet_di_lawan)
                    mutasi_kredit += (0 if sum_mutasi_kredit_di_lawan == None else sum_mutasi_kredit_di_lawan)

            saldo_awal = mutasi_awal_debet - mutasi_awal_kredit
            saldo_akhir = (mutasi_awal_debet + mutasi_debet) - (mutasi_awal_kredit + mutasi_kredit)

            total_mutasi_debet += mutasi_debet
            total_mutasi_kredit += mutasi_kredit

            if item.tipe == 'debit':
                total_saldo_awal_debet += saldo_awal
                total_saldo_akhir_debet += saldo_akhir
            else:
                total_saldo_awal_kredit += saldo_awal
                total_saldo_akhir_kredit += saldo_akhir
            
            saldo_awal_display = saldo_awal
            saldo_akhir_display = saldo_akhir

            if item.tipe == 'kredit':
                saldo_awal_display = saldo_awal * -1
                saldo_akhir_display = saldo_akhir * -1

            saldo_dict = {
                'saldo_awal': saldo_awal,
                'saldo_akhir': saldo_akhir,
                'saldo_awal_display': saldo_awal_display,
                'saldo_akhir_display': saldo_akhir_display,
                'mutasi_debet': mutasi_debet,
                'mutasi_kredit': mutasi_kredit,
            }

            # Add saldo_dict to item
            if isinstance(item, dict):
                item['saldo_dict'] = saldo_dict
            else:
                setattr(item, 'saldo_dict', saldo_dict)

        context['total_saldo_awal_debet'] = total_saldo_awal_debet
        context['total_saldo_awal_kredit'] = total_saldo_awal_kredit * -1
        context['total_mutasi_debet'] = total_mutasi_debet
        context['total_mutasi_kredit'] = total_mutasi_kredit
        context['total_saldo_akhir_debet'] = total_saldo_akhir_debet
        context['total_saldo_akhir_kredit'] = total_saldo_akhir_kredit * -1
    else:
        context['selected_kode_akun'] = []
        context['query_params'] = None
        context['option_selected_kode_akun'] = None

    return render(request, 'general-ledger/neraca-saldo.html', context)

def export_to_excel_neraca(request):
    form = request.GET
    kode_akun = form.getlist('kode_akun', None)
    dari = form.get('dari', None)
    sampai = form.get('sampai', None)
    dari_formated = (date_format(dari, '%Y-%m-%d', '%d-%m-%Y') if dari != None else dari)
    sampai_formated = (date_format(sampai, '%Y-%m-%d', '%d-%m-%Y') if sampai != None else sampai)

    if len(kode_akun) > 0 and dari and sampai:
        # Handle 'all' case if necessary
        if 'all' in kode_akun:
            kode_akun = kode_akun[0]

        if isinstance(kode_akun, list):
            selected_kode_akun = KodeAkun.objects.filter(kode_akun__in=kode_akun)
        else:
            selected_kode_akun = KodeAkun.objects.all()
        context['selected_kode_akun'] = selected_kode_akun

        total_saldo_awal_debet = 0
        total_saldo_awal_kredit = 0
        total_mutasi_debet = 0
        total_mutasi_kredit = 0
        total_saldo_akhir_debet = 0
        total_saldo_akhir_kredit = 0

        for item in selected_kode_akun:
            mutasi_awal_debet = 0
            mutasi_awal_kredit = 0

            mutasi_debet = 0
            mutasi_kredit = 0

            # cek apakah ada jurnal awal di field kode
            cek_transaksi_awal_di_kode = Jurnal.objects.filter(
                kode_akun=item.kode_akun,
                tanggal__lt=dari
            ).exists()
            
            if cek_transaksi_awal_di_kode:
                sum_mutasi_awal_debet_di_kode = Jurnal.objects.filter(
                    kode_akun=item.kode_akun,
                    tanggal__lt=dari,
                    tipe='debit'
                ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                sum_mutasi_awal_kredit_di_kode = Jurnal.objects.filter(
                    kode_akun=item.kode_akun,
                    tanggal__lt=dari,
                    tipe='kredit'
                ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                
                mutasi_awal_debet += sum_mutasi_awal_debet_di_kode
                mutasi_awal_kredit += sum_mutasi_awal_kredit_di_kode

                # cek apakah transaksi sebelumnya juga terdapat di field lawan
                cek_transaksi_lawan_di_lawan = Jurnal.objects.filter(
                                                kode_lawan=item.kode_akun,
                                                tanggal__lt=dari
                                            ).exists()

                if cek_transaksi_lawan_di_lawan:
                    sum_mutasi_awal_debet_di_lawan = Jurnal.objects.filter(
                                                        kode_lawan=item.kode_akun,
                                                        tanggal__lt=dari,
                                                        tipe='debit'
                                                    ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                    sum_mutasi_awal_kredit_di_lawan = Jurnal.objects.filter(
                                                        kode_lawan=item.kode_akun,
                                                        tanggal__lt=dari,
                                                        tipe='kredit'
                                                    ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                    mutasi_awal_debet += sum_mutasi_awal_debet_di_lawan
                    mutasi_awal_kredit += sum_mutasi_awal_kredit_di_lawan
            else:
                # cek apakah ada jurnal awal di field lawan
                cek_transaksi_awal_di_lawan = Jurnal.objects.filter(
                                                    kode_lawan=item.kode_akun,
                                                    tanggal__lt=dari
                                                ).exists()
                if cek_transaksi_awal_di_lawan:
                    sum_mutasi_awal_debet_di_lawan = Jurnal.objects.filter(
                                                        kode_lawan=item.kode_akun,
                                                        tanggal__lt=dari,
                                                        tipe='kredit'
                                                    ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                    sum_mutasi_awal_kredit_di_lawan = Jurnal.objects.filter(
                                                        kode_lawan=item.kode_akun,
                                                        tanggal__lt=dari,
                                                        tipe='debit'
                                                    ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                    mutasi_awal_debet += sum_mutasi_awal_debet_di_lawan
                    mutasi_awal_kredit += sum_mutasi_awal_kredit_di_lawan

            # cek transaksi di field kode
            cek_transaksi_di_kode = Jurnal.objects.filter(
                Q(tanggal__range=(dari,sampai)),
                kode_akun=item.kode_akun
            ).exists()

            if cek_transaksi_di_kode:
                sum_mutasi_debet_di_kode = Jurnal.objects.filter(
                                            Q(tanggal__range=(dari,sampai)),
                                            kode_akun=item.kode_akun,
                                            tipe='debit'
                                        ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                sum_mutasi_kredit_di_kode = Jurnal.objects.filter(
                                            Q(tanggal__range=(dari,sampai)),
                                            kode_akun=item.kode_akun,
                                            tipe='kredit'
                                        ).aggregate(total_nominal=Sum('nominal'))['total_nominal']

                mutasi_debet += (0 if sum_mutasi_debet_di_kode == None else sum_mutasi_debet_di_kode)
                mutasi_kredit += (0 if sum_mutasi_kredit_di_kode == None else sum_mutasi_kredit_di_kode)

                # cek transaksi di field lawan
                cek_transaksi_di_lawan = Jurnal.objects.filter(
                                            Q(tanggal__range=(dari,sampai)),
                                            kode_lawan=item.kode_akun
                                        ).exists()
                if cek_transaksi_di_lawan:
                    sum_mutasi_debet_di_lawan = Jurnal.objects.filter(
                                            Q(tanggal__range=(dari,sampai)),
                                            kode_lawan=item.kode_akun,
                                            tipe='kredit'
                                        ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                    sum_mutasi_kredit_di_lawan = Jurnal.objects.filter(
                                                Q(tanggal__range=(dari,sampai)),
                                                kode_lawan=item.kode_akun,
                                                tipe='debit'
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']

                    mutasi_debet += (0 if sum_mutasi_debet_di_lawan == None else sum_mutasi_debet_di_lawan)
                    mutasi_kredit += (0 if sum_mutasi_kredit_di_lawan == None else sum_mutasi_kredit_di_lawan)
            else:
                # cek transaksi di field lawan
                cek_transaksi_di_lawan = Jurnal.objects.filter(
                                            Q(tanggal__range=(dari,sampai)),
                                            kode_lawan=item.kode_akun
                                        ).exists()
                if cek_transaksi_di_lawan:
                    sum_mutasi_debet_di_lawan = Jurnal.objects.filter(
                                            Q(tanggal__range=(dari,sampai)),
                                            kode_lawan=item.kode_akun,
                                            tipe='kredit'
                                        ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                    sum_mutasi_kredit_di_lawan = Jurnal.objects.filter(
                                                Q(tanggal__range=(dari,sampai)),
                                                kode_lawan=item.kode_akun,
                                                tipe='debit'
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']

                    mutasi_debet += (0 if sum_mutasi_debet_di_lawan == None else sum_mutasi_debet_di_lawan)
                    mutasi_kredit += (0 if sum_mutasi_kredit_di_lawan == None else sum_mutasi_kredit_di_lawan)

            saldo_awal = mutasi_awal_debet - mutasi_awal_kredit
            saldo_akhir = (mutasi_awal_debet + mutasi_debet) - (mutasi_awal_kredit + mutasi_kredit)

            total_mutasi_debet += mutasi_debet
            total_mutasi_kredit += mutasi_kredit

            if item.tipe == 'debit':
                total_saldo_awal_debet += saldo_awal
                total_saldo_akhir_debet += saldo_akhir
            else:
                total_saldo_awal_kredit += saldo_awal
                total_saldo_akhir_kredit += saldo_akhir
            
            saldo_awal_display = saldo_awal
            saldo_akhir_display = saldo_akhir

            if item.tipe == 'kredit':
                saldo_awal_display = saldo_awal * -1
                saldo_akhir_display = saldo_akhir * -1

            saldo_dict = {
                'saldo_awal': saldo_awal,
                'saldo_akhir': saldo_akhir,
                'saldo_awal_display': saldo_awal_display,
                'saldo_akhir_display': saldo_akhir_display,
                'mutasi_debet': mutasi_debet,
                'mutasi_kredit': mutasi_kredit,
            }

            # Add saldo_dict to item
            if isinstance(item, dict):
                item['saldo_dict'] = saldo_dict
            else:
                setattr(item, 'saldo_dict', saldo_dict)

        total_saldo_awal_debet = total_saldo_awal_debet
        total_saldo_awal_kredit = total_saldo_awal_kredit * -1
        total_mutasi_debet = total_mutasi_debet
        total_mutasi_kredit = total_mutasi_kredit
        total_saldo_akhir_debet = total_saldo_akhir_debet
        total_saldo_akhir_kredit = total_saldo_akhir_kredit * -1

    # Create a new Workbook
    filename = f"Neraca Saldo Periode {dari} sd {sampai}"
    wb = Workbook()
    ws = wb.active
    ws.title = filename
    
    # Header row
    header = [
        "Kode Akun", "Nama Akun", "Saldo Awal Debet", "Saldo Awal Kredit", 
        "Mutasi Debet", "Mutasi Kredit", "Saldo Akhir Debet", "Saldo Akhir Kredit"
    ]
    ws.append(header)
    
    # Data rows
    for selected in selected_kode_akun:
        row = [
            selected.kode_akun, selected.nama,
            selected.saldo_dict['saldo_awal'] if selected.tipe == 'debit' else '',
            selected.saldo_dict['saldo_awal'] if selected.tipe == 'kredit' else '',
            selected.saldo_dict['mutasi_debet'],
            selected.saldo_dict['mutasi_kredit'],
            selected.saldo_dict['saldo_akhir'] if selected.tipe == 'debit' else '',
            selected.saldo_dict['saldo_akhir'] if selected.tipe == 'kredit' else '',
        ]
        ws.append(row)
    
    # Save the workbook
    response = HttpResponse(content_type='application/vnd.openpyxl.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
    wb.save(response)

    return response
# END Neraca Saldo

# Laba rugi
def index_laba_rugi(request):
    context['title'] = 'Laba Rugi'
    total_penjualan = 0
    total_beban = 0
    total_pajak = 0
    laba_rugi_kotor = 0
    laba_rugi_sebelum_pajak = 0
    laba_rugi_bersih = 0
    context['laba_rugi'] = None
    form = request.GET
    context['fieldValues'] = form
    bulan = form.get('bulan', None)
    tahun = form.get('tahun', None)
    rekening_penjualan = None
    rekening_beban = None
    rekening_pajak = None

    if bulan != None and tahun != None:
        rekening_penjualan = KodeAkun.objects.filter(
            Q(kode_akun__icontains='4')
        ).order_by('kode_akun')
        rekening_beban = KodeAkun.objects.filter(
            Q(kode_akun__icontains='5')
        ).order_by('kode_akun')
        rekening_pajak = KodeAkun.objects.filter(
            Q(kode_akun__icontains='6')
        ).order_by('kode_akun')

        # rekening penjualan
        for item in rekening_penjualan:
            mutasi_debet = 0
            mutasi_kredit = 0

            cek_transaksi_di_kode = ViewLabaRugi.objects.filter(
                                        bulan=bulan,
                                        tahun=tahun,
                                        kode_akun=item.kode_akun
                                    ).exists()
            if cek_transaksi_di_kode:
                sum_mutasi_debet_kode = ViewLabaRugi.objects.filter(
                                            bulan=bulan,
                                            tahun=tahun,
                                            kode_akun=item.kode_akun,
                                            tipe='debit'
                                        ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                sum_mutasi_kredit_kode = ViewLabaRugi.objects.filter(
                                            bulan=bulan,
                                            tahun=tahun,
                                            kode_akun=item.kode_akun,
                                            tipe='kredit'
                                        ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                
                mutasi_debet += (0 if sum_mutasi_debet_kode == None else sum_mutasi_debet_kode)
                mutasi_kredit += (0 if sum_mutasi_kredit_kode == None else sum_mutasi_kredit_kode)

                cek_transaksi_di_lawan = ViewLabaRugi.objects.filter(
                                        bulan=bulan,
                                        tahun=tahun,
                                        kode_lawan=item.kode_akun
                                    ).exists()
                if cek_transaksi_di_lawan:
                    sum_mutasi_debet_lawan = ViewLabaRugi.objects.filter(
                                                bulan=bulan,
                                                tahun=tahun,
                                                kode_lawan=item.kode_akun,
                                                tipe='kredit'
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                    sum_mutasi_kredit_lawan = ViewLabaRugi.objects.filter(
                                                bulan=bulan,
                                                tahun=tahun,
                                                kode_lawan=item.kode_akun,
                                                tipe='debit'
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']

                    mutasi_debet += (0 if sum_mutasi_debet_lawan == None else sum_mutasi_debet_lawan)
                    mutasi_kredit += (0 if sum_mutasi_kredit_lawan == None else sum_mutasi_kredit_lawan)
            else:
                cek_transaksi_di_lawan = ViewLabaRugi.objects.filter(
                                        bulan=bulan,
                                        tahun=tahun,
                                        kode_lawan=item.kode_akun
                                    ).exists()
                if cek_transaksi_di_lawan:
                    sum_mutasi_debet_lawan = ViewLabaRugi.objects.filter(
                                                bulan=bulan,
                                                tahun=tahun,
                                                kode_lawan=item.kode_akun,
                                                tipe='kredit'
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                    sum_mutasi_kredit_lawan = ViewLabaRugi.objects.filter(
                                                bulan=bulan,
                                                tahun=tahun,
                                                kode_lawan=item.kode_akun,
                                                tipe='debit'
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']

                    mutasi_debet += (0 if sum_mutasi_debet_lawan == None else sum_mutasi_debet_lawan)
                    mutasi_kredit += (0 if sum_mutasi_kredit_lawan == None else sum_mutasi_kredit_lawan)

            penjualan = mutasi_kredit - mutasi_debet
            total_penjualan += penjualan

            if isinstance(item, dict):
                item['penjualan'] = penjualan
            else:
                setattr(item, 'penjualan', penjualan)

        # rekening beban
        for item in rekening_beban:
            mutasi_debet = 0
            mutasi_kredit = 0

            cek_transaksi_di_kode = ViewLabaRugi.objects.filter(
                                        bulan=bulan,
                                        tahun=tahun,
                                        kode_akun=item.kode_akun
                                    ).exists()
            if cek_transaksi_di_kode:
                sum_mutasi_debet_kode = ViewLabaRugi.objects.filter(
                                            bulan=bulan,
                                            tahun=tahun,
                                            kode_akun=item.kode_akun,
                                            tipe='debit'
                                        ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                sum_mutasi_kredit_kode = ViewLabaRugi.objects.filter(
                                            bulan=bulan,
                                            tahun=tahun,
                                            kode_akun=item.kode_akun,
                                            tipe='kredit'
                                        ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                
                mutasi_debet += (0 if sum_mutasi_debet_kode == None else sum_mutasi_debet_kode)
                mutasi_kredit += (0 if sum_mutasi_kredit_kode == None else sum_mutasi_kredit_kode)

                cek_transaksi_di_lawan = ViewLabaRugi.objects.filter(
                                        bulan=bulan,
                                        tahun=tahun,
                                        kode_lawan=item.kode_akun
                                    ).exists()
                if cek_transaksi_di_lawan:
                    sum_mutasi_debet_lawan = ViewLabaRugi.objects.filter(
                                                bulan=bulan,
                                                tahun=tahun,
                                                kode_lawan=item.kode_akun,
                                                tipe='kredit'
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                    sum_mutasi_kredit_lawan = ViewLabaRugi.objects.filter(
                                                bulan=bulan,
                                                tahun=tahun,
                                                kode_lawan=item.kode_akun,
                                                tipe='debit'
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']

                    mutasi_debet += (0 if sum_mutasi_debet_lawan == None else sum_mutasi_debet_lawan)
                    mutasi_kredit += (0 if sum_mutasi_kredit_lawan == None else sum_mutasi_kredit_lawan)
            else:
                cek_transaksi_di_lawan = ViewLabaRugi.objects.filter(
                                        bulan=bulan,
                                        tahun=tahun,
                                        kode_lawan=item.kode_akun
                                    ).exists()
                if cek_transaksi_di_lawan:
                    sum_mutasi_debet_lawan = ViewLabaRugi.objects.filter(
                                                bulan=bulan,
                                                tahun=tahun,
                                                kode_lawan=item.kode_akun,
                                                tipe='kredit'
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                    sum_mutasi_kredit_lawan = ViewLabaRugi.objects.filter(
                                                bulan=bulan,
                                                tahun=tahun,
                                                kode_lawan=item.kode_akun,
                                                tipe='debit'
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']

                    mutasi_debet += (0 if sum_mutasi_debet_lawan == None else sum_mutasi_debet_lawan)
                    mutasi_kredit += (0 if sum_mutasi_kredit_lawan == None else sum_mutasi_kredit_lawan)

            if item.tipe == 'debit':
                beban = mutasi_debet - mutasi_kredit
                total_beban += beban
            else:
                beban = mutasi_kredit - mutasi_debet
                total_beban -= beban

            if isinstance(item, dict):
                item['beban'] = beban
            else:
                setattr(item, 'beban', beban)

        # rekening pajak
        for item in rekening_pajak:
            mutasi_debet = 0
            mutasi_kredit = 0

            cek_transaksi_di_kode = ViewLabaRugi.objects.filter(
                                        bulan=bulan,
                                        tahun=tahun,
                                        kode_akun=item.kode_akun
                                    ).exists()
            if cek_transaksi_di_kode:
                sum_mutasi_debet_kode = ViewLabaRugi.objects.filter(
                                            bulan=bulan,
                                            tahun=tahun,
                                            kode_akun=item.kode_akun,
                                            tipe='debit'
                                        ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                sum_mutasi_kredit_kode = ViewLabaRugi.objects.filter(
                                            bulan=bulan,
                                            tahun=tahun,
                                            kode_akun=item.kode_akun,
                                            tipe='kredit'
                                        ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                
                mutasi_debet += (0 if sum_mutasi_debet_kode == None else sum_mutasi_debet_kode)
                mutasi_kredit += (0 if sum_mutasi_kredit_kode == None else sum_mutasi_kredit_kode)

                cek_transaksi_di_lawan = ViewLabaRugi.objects.filter(
                                        bulan=bulan,
                                        tahun=tahun,
                                        kode_lawan=item.kode_akun
                                    ).exists()
                if cek_transaksi_di_lawan:
                    sum_mutasi_debet_lawan = ViewLabaRugi.objects.filter(
                                                bulan=bulan,
                                                tahun=tahun,
                                                kode_lawan=item.kode_akun,
                                                tipe='kredit'
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                    sum_mutasi_kredit_lawan = ViewLabaRugi.objects.filter(
                                                bulan=bulan,
                                                tahun=tahun,
                                                kode_lawan=item.kode_akun,
                                                tipe='debit'
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']

                    mutasi_debet += (0 if sum_mutasi_debet_lawan == None else sum_mutasi_debet_lawan)
                    mutasi_kredit += (0 if sum_mutasi_kredit_lawan == None else sum_mutasi_kredit_lawan)
            else:
                cek_transaksi_di_lawan = ViewLabaRugi.objects.filter(
                                        bulan=bulan,
                                        tahun=tahun,
                                        kode_lawan=item.kode_akun
                                    ).exists()
                if cek_transaksi_di_lawan:
                    sum_mutasi_debet_lawan = ViewLabaRugi.objects.filter(
                                                bulan=bulan,
                                                tahun=tahun,
                                                kode_lawan=item.kode_akun,
                                                tipe='kredit'
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                    sum_mutasi_kredit_lawan = ViewLabaRugi.objects.filter(
                                                bulan=bulan,
                                                tahun=tahun,
                                                kode_lawan=item.kode_akun,
                                                tipe='debit'
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']

                    mutasi_debet += (0 if sum_mutasi_debet_lawan == None else sum_mutasi_debet_lawan)
                    mutasi_kredit += (0 if sum_mutasi_kredit_lawan == None else sum_mutasi_kredit_lawan)

            if item.tipe == 'debit':
                pajak = mutasi_debet - mutasi_kredit
                total_pajak += pajak
            else:
                pajak = mutasi_kredit - mutasi_debet
                total_pajak -= pajak

            if isinstance(item, dict):
                item['pajak'] = pajak
            else:
                setattr(item, 'pajak', pajak)

    laba_rugi_kotor = total_penjualan
    laba_rugi_sebelum_pajak = laba_rugi_kotor - total_beban
    laba_rugi_bersih = laba_rugi_sebelum_pajak - total_pajak
    context['total_penjualan'] = total_penjualan
    context['total_beban'] = total_beban
    context['total_pajak'] = total_pajak
    context['laba_rugi_kotor'] = laba_rugi_kotor
    context['laba_rugi_sebelum_pajak'] = laba_rugi_sebelum_pajak
    context['laba_rugi_bersih'] = laba_rugi_bersih

    context['rekening_penjualan'] = rekening_penjualan
    context['rekening_beban'] = rekening_beban
    context['rekening_pajak'] = rekening_pajak

    return render(request, 'general-ledger/laba-rugi.html', context)

def export_to_excel_laba_rugi(request):
    total_penjualan = 0
    total_beban = 0
    total_pajak = 0
    laba_rugi_kotor = 0
    laba_rugi_sebelum_pajak = 0
    laba_rugi_bersih = 0
    context['laba_rugi'] = None
    form = request.GET
    context['fieldValues'] = form
    bulan = form.get('bulan', None)
    tahun = form.get('tahun', None)
    rekening_penjualan = []
    rekening_beban = []
    rekening_pajak = []

    if bulan != None and tahun != None:
        rekening_penjualan = KodeAkun.objects.filter(
            Q(kode_akun__icontains='4')
        ).order_by('kode_akun')
        rekening_beban = KodeAkun.objects.filter(
            Q(kode_akun__icontains='5')
        ).order_by('kode_akun')
        rekening_pajak = KodeAkun.objects.filter(
            Q(kode_akun__icontains='6')
        ).order_by('kode_akun')

        # rekening penjualan
        for item in rekening_penjualan:
            mutasi_debet = 0
            mutasi_kredit = 0

            cek_transaksi_di_kode = ViewLabaRugi.objects.filter(
                                        bulan=bulan,
                                        tahun=tahun,
                                        kode_akun=item.kode_akun
                                    ).exists()
            if cek_transaksi_di_kode:
                sum_mutasi_debet_kode = ViewLabaRugi.objects.filter(
                                            bulan=bulan,
                                            tahun=tahun,
                                            kode_akun=item.kode_akun,
                                            tipe='debit'
                                        ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                sum_mutasi_kredit_kode = ViewLabaRugi.objects.filter(
                                            bulan=bulan,
                                            tahun=tahun,
                                            kode_akun=item.kode_akun,
                                            tipe='kredit'
                                        ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                
                mutasi_debet += (0 if sum_mutasi_debet_kode == None else sum_mutasi_debet_kode)
                mutasi_kredit += (0 if sum_mutasi_kredit_kode == None else sum_mutasi_kredit_kode)

                cek_transaksi_di_lawan = ViewLabaRugi.objects.filter(
                                        bulan=bulan,
                                        tahun=tahun,
                                        kode_lawan=item.kode_akun
                                    ).exists()
                if cek_transaksi_di_lawan:
                    sum_mutasi_debet_lawan = ViewLabaRugi.objects.filter(
                                                bulan=bulan,
                                                tahun=tahun,
                                                kode_lawan=item.kode_akun,
                                                tipe='kredit'
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                    sum_mutasi_kredit_lawan = ViewLabaRugi.objects.filter(
                                                bulan=bulan,
                                                tahun=tahun,
                                                kode_lawan=item.kode_akun,
                                                tipe='debit'
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']

                    mutasi_debet += (0 if sum_mutasi_debet_lawan == None else sum_mutasi_debet_lawan)
                    mutasi_kredit += (0 if sum_mutasi_kredit_lawan == None else sum_mutasi_kredit_lawan)
            else:
                cek_transaksi_di_lawan = ViewLabaRugi.objects.filter(
                                        bulan=bulan,
                                        tahun=tahun,
                                        kode_lawan=item.kode_akun
                                    ).exists()
                if cek_transaksi_di_lawan:
                    sum_mutasi_debet_lawan = ViewLabaRugi.objects.filter(
                                                bulan=bulan,
                                                tahun=tahun,
                                                kode_lawan=item.kode_akun,
                                                tipe='kredit'
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                    sum_mutasi_kredit_lawan = ViewLabaRugi.objects.filter(
                                                bulan=bulan,
                                                tahun=tahun,
                                                kode_lawan=item.kode_akun,
                                                tipe='debit'
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']

                    mutasi_debet += (0 if sum_mutasi_debet_lawan == None else sum_mutasi_debet_lawan)
                    mutasi_kredit += (0 if sum_mutasi_kredit_lawan == None else sum_mutasi_kredit_lawan)

            penjualan = mutasi_kredit - mutasi_debet
            total_penjualan += penjualan

            if isinstance(item, dict):
                item['penjualan'] = penjualan
            else:
                setattr(item, 'penjualan', penjualan)

        # rekening beban
        for item in rekening_beban:
            mutasi_debet = 0
            mutasi_kredit = 0

            cek_transaksi_di_kode = ViewLabaRugi.objects.filter(
                                        bulan=bulan,
                                        tahun=tahun,
                                        kode_akun=item.kode_akun
                                    ).exists()
            if cek_transaksi_di_kode:
                sum_mutasi_debet_kode = ViewLabaRugi.objects.filter(
                                            bulan=bulan,
                                            tahun=tahun,
                                            kode_akun=item.kode_akun,
                                            tipe='debit'
                                        ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                sum_mutasi_kredit_kode = ViewLabaRugi.objects.filter(
                                            bulan=bulan,
                                            tahun=tahun,
                                            kode_akun=item.kode_akun,
                                            tipe='kredit'
                                        ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                
                mutasi_debet += (0 if sum_mutasi_debet_kode == None else sum_mutasi_debet_kode)
                mutasi_kredit += (0 if sum_mutasi_kredit_kode == None else sum_mutasi_kredit_kode)

                cek_transaksi_di_lawan = ViewLabaRugi.objects.filter(
                                        bulan=bulan,
                                        tahun=tahun,
                                        kode_lawan=item.kode_akun
                                    ).exists()
                if cek_transaksi_di_lawan:
                    sum_mutasi_debet_lawan = ViewLabaRugi.objects.filter(
                                                bulan=bulan,
                                                tahun=tahun,
                                                kode_lawan=item.kode_akun,
                                                tipe='kredit'
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                    sum_mutasi_kredit_lawan = ViewLabaRugi.objects.filter(
                                                bulan=bulan,
                                                tahun=tahun,
                                                kode_lawan=item.kode_akun,
                                                tipe='debit'
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']

                    mutasi_debet += (0 if sum_mutasi_debet_lawan == None else sum_mutasi_debet_lawan)
                    mutasi_kredit += (0 if sum_mutasi_kredit_lawan == None else sum_mutasi_kredit_lawan)
            else:
                cek_transaksi_di_lawan = ViewLabaRugi.objects.filter(
                                        bulan=bulan,
                                        tahun=tahun,
                                        kode_lawan=item.kode_akun
                                    ).exists()
                if cek_transaksi_di_lawan:
                    sum_mutasi_debet_lawan = ViewLabaRugi.objects.filter(
                                                bulan=bulan,
                                                tahun=tahun,
                                                kode_lawan=item.kode_akun,
                                                tipe='kredit'
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                    sum_mutasi_kredit_lawan = ViewLabaRugi.objects.filter(
                                                bulan=bulan,
                                                tahun=tahun,
                                                kode_lawan=item.kode_akun,
                                                tipe='debit'
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']

                    mutasi_debet += (0 if sum_mutasi_debet_lawan == None else sum_mutasi_debet_lawan)
                    mutasi_kredit += (0 if sum_mutasi_kredit_lawan == None else sum_mutasi_kredit_lawan)

            if item.tipe == 'debit':
                beban = mutasi_debet - mutasi_kredit
                total_beban += beban
            else:
                beban = mutasi_kredit - mutasi_debet
                total_beban -= beban

            if isinstance(item, dict):
                item['beban'] = beban
            else:
                setattr(item, 'beban', beban)

        # rekening pajak
        for item in rekening_pajak:
            mutasi_debet = 0
            mutasi_kredit = 0

            cek_transaksi_di_kode = ViewLabaRugi.objects.filter(
                                        bulan=bulan,
                                        tahun=tahun,
                                        kode_akun=item.kode_akun
                                    ).exists()
            if cek_transaksi_di_kode:
                sum_mutasi_debet_kode = ViewLabaRugi.objects.filter(
                                            bulan=bulan,
                                            tahun=tahun,
                                            kode_akun=item.kode_akun,
                                            tipe='debit'
                                        ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                sum_mutasi_kredit_kode = ViewLabaRugi.objects.filter(
                                            bulan=bulan,
                                            tahun=tahun,
                                            kode_akun=item.kode_akun,
                                            tipe='kredit'
                                        ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                
                mutasi_debet += (0 if sum_mutasi_debet_kode == None else sum_mutasi_debet_kode)
                mutasi_kredit += (0 if sum_mutasi_kredit_kode == None else sum_mutasi_kredit_kode)

                cek_transaksi_di_lawan = ViewLabaRugi.objects.filter(
                                        bulan=bulan,
                                        tahun=tahun,
                                        kode_lawan=item.kode_akun
                                    ).exists()
                if cek_transaksi_di_lawan:
                    sum_mutasi_debet_lawan = ViewLabaRugi.objects.filter(
                                                bulan=bulan,
                                                tahun=tahun,
                                                kode_lawan=item.kode_akun,
                                                tipe='kredit'
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                    sum_mutasi_kredit_lawan = ViewLabaRugi.objects.filter(
                                                bulan=bulan,
                                                tahun=tahun,
                                                kode_lawan=item.kode_akun,
                                                tipe='debit'
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']

                    mutasi_debet += (0 if sum_mutasi_debet_lawan == None else sum_mutasi_debet_lawan)
                    mutasi_kredit += (0 if sum_mutasi_kredit_lawan == None else sum_mutasi_kredit_lawan)
            else:
                cek_transaksi_di_lawan = ViewLabaRugi.objects.filter(
                                        bulan=bulan,
                                        tahun=tahun,
                                        kode_lawan=item.kode_akun
                                    ).exists()
                if cek_transaksi_di_lawan:
                    sum_mutasi_debet_lawan = ViewLabaRugi.objects.filter(
                                                bulan=bulan,
                                                tahun=tahun,
                                                kode_lawan=item.kode_akun,
                                                tipe='kredit'
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']
                    sum_mutasi_kredit_lawan = ViewLabaRugi.objects.filter(
                                                bulan=bulan,
                                                tahun=tahun,
                                                kode_lawan=item.kode_akun,
                                                tipe='debit'
                                            ).aggregate(total_nominal=Sum('nominal'))['total_nominal']

                    mutasi_debet += (0 if sum_mutasi_debet_lawan == None else sum_mutasi_debet_lawan)
                    mutasi_kredit += (0 if sum_mutasi_kredit_lawan == None else sum_mutasi_kredit_lawan)

            if item.tipe == 'debit':
                pajak = mutasi_debet - mutasi_kredit
                total_pajak += pajak
            else:
                pajak = mutasi_kredit - mutasi_debet
                total_pajak -= pajak

            if isinstance(item, dict):
                item['pajak'] = pajak
            else:
                setattr(item, 'pajak', pajak)

        laba_rugi_kotor = total_penjualan
        laba_rugi_sebelum_pajak = laba_rugi_kotor - total_beban
        laba_rugi_bersih = laba_rugi_sebelum_pajak - total_pajak

    # Create a new Workbook
    filename = f"Laba Rugi Periode {bulan} - {tahun}"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename

    # Write the header row
    headers = [
        filename
    ]
    bold_font = Font(bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = bold_font

    # Adding pendapatan section
    ws.cell(row=2, column=1, value='Pendapatan')

    row_num = 3
    for item in rekening_penjualan:
        ws.cell(row=row_num, column=1, value=item.nama)
        ws.cell(row=row_num, column=2, value=item.penjualan)
        row_num += 1
    
    ws.cell(row=row_num, column=1, value='Total Pendapatan')
    ws.cell(row=row_num, column=2, value=total_penjualan)
    laba_rugi_kotor = total_penjualan  # Assuming laba_rugi_kotor calculation
    row_num += 1

    ws.cell(row=row_num, column=1, value='Laba Rugi Kotor')
    ws.cell(row=row_num, column=2, value=laba_rugi_kotor)
    row_num += 1

    for item in rekening_beban:
        ws.cell(row=row_num, column=1, value=item.nama)
        ws.cell(row=row_num, column=2, value=item.beban)
        row_num += 1

    ws.cell(row=row_num, column=1, value='Total Beban')
    ws.cell(row=row_num, column=2, value=total_beban)
    laba_rugi_sebelum_pajak = laba_rugi_kotor - total_beban  # Assuming laba_rugi_sebelum_pajak calculation
    row_num += 1

    ws.cell(row=row_num, column=1, value='Laba Rugi Sebelum Pajak')
    ws.cell(row=row_num, column=2, value=laba_rugi_sebelum_pajak)
    row_num += 1

    for item in rekening_pajak:
        ws.cell(row=row_num, column=1, value=item.nama)
        ws.cell(row=row_num, column=2, value=item.pajak if item.tipe == 'debit' else f'({item.pajak})')
        row_num += 1

    ws.cell(row=row_num, column=1, value='Total Pajak')
    ws.cell(row=row_num, column=2, value=total_pajak)
    row_num += 1

    ws.cell(row=row_num, column=1, value='Laba Rugi Bersih')
    ws.cell(row=row_num, column=2, value=laba_rugi_bersih)

    # Set the column widths
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    # Create an HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)

    return response
# END Laba rugi